#!/usr/bin/env python
"""Fetch all works for an author from OpenAlex by ORCID and save them.

Writes two files into data/:
  - openalex_works.json  (full normalized data, one record per work)
  - openalex_works.xlsx  (spreadsheet for manual review, with a
    "Not mine? (mark X)" column for flagging mismatches)

Usage:
  python -X utf8 scripts/fetch_openalex_works.py [--orcid 0000-0003-1922-439X]

Re-run any time to refresh the data with current citation counts.
"""

import argparse
import json
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

API_BASE = "https://api.openalex.org/works"
DEFAULT_ORCID = "0000-0003-1922-439X"
CONTACT_EMAIL = "vagisha@gmail.com"  # OpenAlex "polite pool" contact, for higher rate limits
PER_PAGE = 200
MAX_RETRIES = 6

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"


def fetch_page(session, orcid, cursor):
    params = {
        "filter": f"author.orcid:{orcid}",
        "per-page": PER_PAGE,
        "cursor": cursor,
        "mailto": CONTACT_EMAIL,
    }
    for attempt in range(1, MAX_RETRIES + 1):
        resp = session.get(API_BASE, params=params, timeout=30)
        if resp.status_code == 429:
            wait = int(resp.headers.get("Retry-After", 2 ** attempt))
            print(f"  Rate limited (429) — waiting {wait}s before retrying "
                  f"(attempt {attempt}/{MAX_RETRIES})...", file=sys.stderr)
            time.sleep(wait)
            continue
        if resp.status_code >= 500:
            wait = 2 ** attempt
            print(f"  OpenAlex server error ({resp.status_code}) — waiting "
                  f"{wait}s before retrying (attempt {attempt}/{MAX_RETRIES})...",
                  file=sys.stderr)
            time.sleep(wait)
            continue
        resp.raise_for_status()
        return resp.json()
    raise RuntimeError(f"Gave up after {MAX_RETRIES} retries fetching {resp.url}")


def best_link(work):
    doi = work.get("doi")
    if doi:
        return doi
    loc = work.get("primary_location") or {}
    if loc.get("landing_page_url"):
        return loc["landing_page_url"]
    return work.get("id")


def venue_name(work):
    loc = work.get("primary_location") or {}
    source = loc.get("source") or {}
    return source.get("display_name") or ""


def author_names(work):
    return [
        a["author"]["display_name"]
        for a in work.get("authorships", [])
        if a.get("author", {}).get("display_name")
    ]


def format_authors_display(names):
    """First and last author, plus up to 4 authors in between."""
    if len(names) <= 6:
        return ", ".join(names)
    middle = names[1:-1][:4]
    return ", ".join([names[0], *middle, "...", names[-1]])


def normalize(work):
    counts_by_year = sorted(
        (
            {"year": c["year"], "cited_by_count": c["cited_by_count"]}
            for c in work.get("counts_by_year", [])
        ),
        key=lambda c: c["year"],
    )
    authors = author_names(work)
    return {
        "openalex_id": work.get("id"),
        "title": work.get("title") or work.get("display_name"),
        "authors": authors,
        "authors_display": format_authors_display(authors),
        "year": work.get("publication_year"),
        "type": work.get("type"),
        "venue": venue_name(work),
        "link": best_link(work),
        "cited_by_count_total": work.get("cited_by_count", 0),
        "citations_by_year": counts_by_year,
    }


def fetch_all_works(orcid):
    session = requests.Session()
    works = []
    cursor = "*"
    page_num = 1
    while cursor:
        print(f"Fetching page {page_num}...", file=sys.stderr)
        data = fetch_page(session, orcid, cursor)
        results = data.get("results", [])
        works.extend(normalize(w) for w in results)
        cursor = (data.get("meta") or {}).get("next_cursor")
        if not results:
            break
        page_num += 1
    return works


def write_json(works, path):
    path.write_text(json.dumps(works, indent=2, ensure_ascii=False), encoding="utf-8")


def write_xlsx(works, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "OpenAlex works"

    headers = [
        "Title", "Authors", "Year", "Type", "Venue", "Link",
        "Total citations", "Citations by year", "Not mine? (mark X)",
    ]
    ws.append(headers)

    for w in works:
        citations_by_year_str = "; ".join(
            f"{c['year']}: {c['cited_by_count']}" for c in w["citations_by_year"]
        )
        ws.append([
            w["title"],
            w["authors_display"],
            w["year"],
            w["type"],
            w["venue"],
            w["link"],
            w["cited_by_count_total"],
            citations_by_year_str,
            "",
        ])

    widths = [55, 40, 8, 14, 30, 45, 12, 40, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    wb.save(path)


def summarize(works):
    by_type = {}
    for w in works:
        by_type[w["type"]] = by_type.get(w["type"], 0) + 1
    print(f"\nFetched {len(works)} works from OpenAlex.")
    print("Breakdown by type:")
    for t, count in sorted(by_type.items(), key=lambda kv: -kv[1]):
        print(f"  {t}: {count}")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--orcid", default=DEFAULT_ORCID)
    args = parser.parse_args()

    DATA_DIR.mkdir(exist_ok=True)
    works = fetch_all_works(args.orcid)
    works.sort(key=lambda w: (w["year"] or 0), reverse=True)

    write_json(works, DATA_DIR / "openalex_works.json")
    write_xlsx(works, DATA_DIR / "openalex_works.xlsx")

    summarize(works)
    print(f"\nSaved: {DATA_DIR / 'openalex_works.json'}")
    print(f"Saved: {DATA_DIR / 'openalex_works.xlsx'}")


if __name__ == "__main__":
    main()
